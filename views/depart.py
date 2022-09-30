from django.shortcuts import render, redirect,HttpResponse
from app01 import models
from django import forms
from django.core.validators import RegexValidator, ValidationError
from django.utils.safestring import mark_safe
from app01.utils.pageaction import PageInaction
from app01.utils.bootstrap import BootStrapModeForm
from app01.utils.form import *

def depart_list(request):
    info=request.session.get("info")
    if not info:
        return redirect("/login/")
    queryset=models.Department.objects.all()
    page_obj=PageInaction(request,queryset,page_size=10)
    context={
        "queryset": page_obj.page_queryset,
        "page_string": page_obj.html()
    }
    return render(request,"depart_list.html",context)


def depart_add(request):
    queryset=models.Department.objects.all()
    if request.method == "GET":
        return render(request,"depart_add.html")
    title=request.POST.get("title")
    models.Department.objects.create(title=title)
    return redirect("/depart/list/")

def depart_delete(request):
    nid=request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")

def depart_edit(request,nid):
    if request.method == "GET":
        row_data = models.Department.objects.filter(id=nid).first()
        return render(request,"depart_edit.html",{"row_object":row_data})

    title=request.POST.get("title")
    models.Department.objects.filter(id=nid).update(title=title)
    #print(row_data.id, row_data.title)
    return redirect("/depart/list/")

def depart_multi(request):
    from openpyxl import load_workbook
    file_object=request.FILES.get("exc")
    f = open(file_object.name,mode='wb')

    wb =load_workbook(file_object)
    sheet=wb.worksheets[0]
    cell =sheet.cell(1,1)
    print(cell.value)
    for row in sheet.iter_rows(min_row=2):
        text=row[0].value
        print(text)
        exits = models.Department.objects.filter(title=text).exists()
        if not exits:
            models.Department.objects.create(title=text)
    return redirect("/depart/list/")

