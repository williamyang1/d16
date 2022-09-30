from django.shortcuts import render, redirect,HttpResponse
from app01 import models
from django import forms
from django.core.validators import RegexValidator, ValidationError
from django.utils.safestring import mark_safe
from app01.utils.pageaction import PageInaction
from app01.utils.bootstrap import BootStrapModeForm
from app01.utils.form import *
from django.db.models import Q
from app01.utils.get_bug_data_from_nvbug import NvbugsUtils

def bugs_update(request):

    bug = NvbugsUtils("williamy", "Williamy1203#")
    print(bug.get_bug_details(200742297))

    cuDNN_QA_filed_open_Bugs = 200045751
    QA_actionable_buts = 200053667
    QA_actionable_buts=[3785542, 3782813,3782592,3775740]
    for bugid in QA_actionable_buts:
        if models.NvBug.objects.filter(BugId=bugid).exists():
            print("Exist")

        else:
            print("Not exist")
            BugId,Synopsis,BugAction,Engineer,Version,Module,Disposition, Regression,Keywords,Created,DaysOpen,Origin,Version, ModifiedDate,Priority,RequestDate,Categories,QAEngineer=bug.get_bug_details(bugid)
            buglink="https://nvbugs/"+str(bugid)
            models.NvBug.objects.create(
                BugId=BugId,
                Synopsis=Synopsis,
                BugAction=BugAction,
                Module=Module,
                Priority=Priority,
                RequestDate=RequestDate,
                Categories=Categories,
                Disposition=Disposition,
                QAEngineer=QAEngineer,
                Engineer=Engineer,
                CustomKeywords=Keywords,
                ModifiedDate=ModifiedDate,
                Version=Version,
                Origin=Origin,
                Regression=Regression,
                buglink=buglink,
                DaysOpen=DaysOpen,
            )




    return HttpResponse("Updating")

def bugs_list(request):
    #page=int(request.GET.get("page",1))
    data_dict = {}
    search_data=request.GET.get('q',"")
    if search_data:
        data_dict["CustomKeywords__contains"]=search_data
        data_dict["BugId__contains"]=search_data
    print(data_dict)
    queryset=models.NvBug.objects.filter(Q(CustomKeywords__contains=search_data)|Q(BugId__contains=search_data)).order_by("-BugId")

    page_object = PageInaction(request, queryset,page_size=30)
    page_queryset=page_object.page_queryset
    page_str=page_object.html()
    content={"queryset":page_queryset,
             "search_data":search_data,
             "page_string":page_str}
    return render(request,"bugs_list.html",content)


def bugs_multi(request):
    from openpyxl import load_workbook
    file_object=request.FILES.get("exc")
    f = open(file_object.name,mode='wb')
    print("AAAAAAAAAAA")
    print(file_object)
    wb =load_workbook(file_object)
    sheet=wb.worksheets[0]
    cell =sheet.cell(1,1)
    print(cell.value)

    for row in sheet.iter_rows(min_row=2):

        BugId = row[0].value
        Synopsis = row[1].value
        BugAction = row[2].value
        Module = row[3].value
        Priority = row[4].value
        RequestDate = row[5].value
        Categories = row[6].value
        Disposition = row[7].value
        QAEngineer = row[8].value
        Engineer = row[9].value
        CustomKeywords = row[10].value
        ModifiedDate = row[11].value
        Version = row[12].value
        Origin = row[13].value
        Regression = row[14].value
        Error = row[15].value
        Tlist = row[16].value
        buglink = row[17].value
        DaysOpen = row[18].value

        exits = models.NvBug.objects.filter(BugId=BugId).exists()
        if not exits:
             models.NvBug.objects.create(
                 BugId=BugId,
                 Synopsis=Synopsis,
                 BugAction=BugAction,
                 Module=Module,
                 Priority=Priority,
                 RequestDate=RequestDate,
                 Categories=Categories,
                 Disposition=Disposition,
                 QAEngineer=QAEngineer,
                 Engineer=Engineer,
                 CustomKeywords=CustomKeywords,
                 ModifiedDate=ModifiedDate,
                 Version=Version,
                 Origin=Origin,
                 Regression=Regression,
                 Error=Error,
                 Tlist=Tlist,
                 buglink=buglink,
                 DaysOpen=DaysOpen,
             )
    return redirect("/bugs/list/")